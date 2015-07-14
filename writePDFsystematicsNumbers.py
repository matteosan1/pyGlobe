#!/usr/bin/env python

# calculates the impact of varying the PDFs for the signal
import sys, re, os
import utils
from pprint import pprint
import math

#----------------------------------------------------------------------
wsname = "CMS_emu_workspace"

# determine the systematics at one mass point only
mass = 125

# whether or not to normalize the relative deviations
# such that the total number of expected
# events stays the same. This should be done if one
# includes pdf uncertainties on the cross sections
# in addition in combine, otherwise the
# pdf (normalization) uncertainty is double counted
averageDeviations = True


pdfMembers = [
    ('CT10',            53),
    ('MSTW2008nlo68cl', 41),
    ('NNPDF10_100',    101),
    ]

pdfsWithUpDown = ('CT10', 'MSTW2008nlo68cl')

numPDFs = sum([ line[1] for line in pdfMembers ])
numPDFfamilies = len(pdfMembers)

allCats  = None
allProcs = None

# allCats = [ 'cat0' ]
# allProcs = [ 'ggh' ]

#----------------------------------------------------------------------

def getSumOfWeightsGlobeTuples(proc):
    filePattern = "/home/users/aholz/hadoop/2013-07-hemu/00-crab-globe/sig-with-pdfweights/" + proc + "/125.0/*.root"

    # for the moment, we just take the pdf weights into account,
    # no correlation with the PU reweighting

    import ROOT
    chain = ROOT.TChain("event")
    chain.Add(filePattern)
    
    # get sum of event weights

    retval = [ 0. ] * numPDFs

    numEvents = chain.GetEntries()
    for i in range(numEvents):
        chain.GetEntry(i)

        # get event weight for each pdf
        for j in range(numPDFs):
            weight = chain.weight * chain.pdf_weights[j] / chain.pdf_weights[0]

            retval[j] += weight

    return retval

#----------------------------------------------------------------------

def getSumOfWeightsWorkspace(fname):

    retval = {}

    import ROOT
    fin = ROOT.TFile(fname)
    assert fin.IsOpen(), "could not open file " + fname

    ws = fin.Get(wsname)
    assert ws != None, "could not find workspace " + wsname

    # get list of categories from the first file

    global allCats, allProcs

    if allCats == None:
        allCats   = utils.getCatEntries(utils.getObj(ws, 'allCategories'))

    if allProcs == None:
        allProcs  = utils.getCatEntries(utils.getObj(ws, 'allSigProcesses'))

    for cat in allCats:
        for proc in allProcs:

            # e.g. sig_Hem_unbinned_vbf_120_cat3
            name = "_".join([
                "sig",
                "Hem",
                "unbinned",
                proc,
                str(mass),
                cat
                ])

            # binned dataset
            # e.g. sig_Hem_vbf_115_cat10
            name = "_".join([
                "sig",
                "Hem",
                proc,
                str(mass),
                cat
                ])

            # get the signal MC dataset
            ds = utils.getObj(ws, name)

            retval.setdefault(cat,{})[proc] = ds.sumEntries()

        # end of loop over signal processes
    # end of loop over categories
    
    return retval

#----------------------------------------------------------------------

# def addSumEventsAfterSelectionFields(topleft):

def addPerPDFfamilyUpDownSummary(row, cellsForMax):
    # adds cells taking the maximum or RMS for the sets in a PDF

    global allCats
    global firstRatioColumn

    # add the sqrt(sumsq(..)) cells
    for catIndex, cat in enumerate(allCats):

        # for number of events
        # col = firstNumEventsColumn + numCats + 1 + 2 * catIndex
        # for 'acceptance'
        col = firstRatioColumn + numCats + 1 + 2 * catIndex

        origCol = openpyxl.cell.get_column_letter(firstRatioColumn + catIndex)

        for plusMinusIndex in range(2):
            colName = openpyxl.cell.get_column_letter(col + plusMinusIndex)

            cell = ws.cell(row = row + familySize + 1,
                           column = col + plusMinusIndex,

                           # relative uncertainty per PDF family
                           # value = '=SQRT(SUMSQ(%s%d:%s%d)) / %s%d' % (colName, row,
                           #                                             colName, row + familySize - 1,
                           #                                             origCol, row)

                           # absolute uncertainty per PDF family, normalize later
                           value = '=SQRT(SUMSQ(%s%d:%s%d))' % (colName, row,
                                                                colName, row + familySize - 1,
                                                                )
                           )

            # only do this for relative uncertainty 
            # cell.number_format = '0.00%'

            cellsForMax.setdefault(cat,[]).append(cell)

            ## cell = ws.cell(row = row + familySize + 1,
            ##                column = col + plusMinusIndex,
            ##                value = '=SQRT(SUMSQ(%s%d:%s%d))' % (colName, row,
            ##                                                            colName, row + familySize - 1,
            ##                                                            )
            ##                )

#----------------------------------------------------------------------

def addPerPDFfamilyNNPDFsummary(row, cellsForMax):

    global allCats
    global firstRatioColumn

    # add the stdev(..) cells
    for catIndex, cat in enumerate(allCats):

        # for number of events
        # col = firstNumEventsColumn + numCats + 1 + 2 * catIndex
        # for 'acceptance'
        col = firstRatioColumn + numCats + 1 + 2 * catIndex

        colName = openpyxl.cell.get_column_letter(firstRatioColumn + catIndex)

        # the range over which we take the average and stddev.
        rangeDesc = "%s%d:%s%d" % (colName, row + 1, # do not include nominal row
                                   colName, row + familySize - 1)

        cell = ws.cell(row = row + familySize + 1,
                column = col,

                # relative uncertainty per PDF family       
                # value = '=STDEV(%s) / AVERAGE(%s)'  % (rangeDesc, rangeDesc),

                # absolute uncertainty per PDF family, normalize later
                 value = '=STDEV(%s)'  % rangeDesc,
                )

        # only do this for relative uncertainty 
        # cell.number_format = '0.00%'
        cellsForMax.setdefault(cat,[]).append(cell)


#----------------------------------------------------------------------

def addMaxOverPDFfamilies(row, nominalRow, cellsForMax):
    # @param nominalCell is the cell with the nominal value
    
    ws.cell(row = row,
            column = 3,
            value = "max over all PDF families")

    # nominalCol = firstNominalCell.column
    # nominalRow = firstNominalCell.row

    for catIndex, cat in enumerate(allCats):
        # column = firstNumEventsColumn + catIndex
        column = firstRatioColumn + catIndex

        nominalCell = ws.cell(column = column,
                              row = nominalRow)

        cellNames = [ cell.coordinate for cell in cellsForMax[cat] ]

        cell = ws.cell(row = row,
                       column = column,
                       value = "=MAX(%s) / %s" % (",".join(cellNames), nominalCell.coordinate)
                       )

        cell.number_format = '0.00%'

        # produce the same cell with 'absolute' format
        # so we can put them into the python files
        # to generate the datacards for combine
        cell = ws.cell(row = row + 1,
                       column = column,
                       value = "=MAX(%s) / %s" % (",".join(cellNames), nominalCell.coordinate)
                       )
        

    



#----------------------------------------------------------------------
# main
#----------------------------------------------------------------------

if False:
    print getSumOfWeightsGlobeTuples('ggh')
    print "--------------------"
    print getSumOfWeightsGlobeTuples('vbf')
    
    sys.exit(1)
else:
    # note that this is independent of (selection) categories
    sumWeightsGlobeTuples = {
        "ggh": [8940.0, 8943.884810978698, 8936.292024288045, 8954.593220015773, 8924.92045992629, 9113.264990274265, 8789.622200304719, 9010.308576524212, 8866.754612266284, 8975.67990133543, 8893.100957959225, 8940.533446594585, 8939.622884233233, 8892.815828767913, 9016.3078201992, 8907.895448598692, 8969.124499671834, 8804.293323728589, 9018.592255493855, 8888.868975301317, 8965.245432264317, 8985.12670699006, 8908.074636895766, 9053.297119203167, 8810.743222430172, 8936.299276944703, 8947.456118143748, 8957.341330116633, 8926.279924470186, 8904.740204992102, 8966.620371235364, 8979.513643512733, 8903.893581600572, 8950.330213547804, 8904.649073784145, 8953.70041221985, 8970.143280314529, 8943.928553945723, 8935.35482758048, 8951.043335216957, 8954.3080061586, 8975.065811892317, 8907.924498308785, 8954.054415081668, 8924.618491102054, 8905.542360743299, 8989.298167305393, 8982.517638792357, 8833.884406491929, 8938.819723170933, 8910.219065206968, 8977.421467674338, 8812.30525327015, 8972.114105845341, 8987.874955481213, 8959.152969245308, 8981.598437051509, 8961.568537243638, 8982.584289940974, 8956.627094822436, 8980.163527574237, 8966.726023599456, 8955.091918403865, 8996.370065103674, 8874.940612439448, 9032.4683852144, 8997.706872643676, 8937.258608579757, 8938.389476053728, 8984.43924671866, 8956.20318175159, 8965.181915328996, 8964.318847163284, 8979.279941736071, 8998.998443008415, 8921.11147658558, 8958.869430100334, 8980.005412340772, 8971.122886331374, 8971.878821780594, 8969.464777134375, 8987.497748932114, 8928.930060236704, 9053.36822082138, 8967.913101367865, 8973.78412256078, 8983.861015023786, 8965.329627888574, 8966.92559064389, 8973.995818930518, 9010.630319841206, 8894.642188943713, 8961.268262481113, 8975.122493538249, 9195.138288302962, 9236.92761223611, 9165.517302560405, 8974.032625745696, 9312.367728282497, 9344.984270522047, 9263.805145971608, 9224.344751319402, 9187.43285650176, 9080.852902882672, 9162.603261608952, 9035.949788373933, 9219.082732764922, 9118.626231885082, 9084.2375700817, 9256.563121355857, 9216.488367453214, 9299.270832509796, 9252.002607046492, 9227.650487996063, 9256.229326315208, 9055.849566968278, 9084.259604489818, 9193.858160879969, 9078.966035860893, 9112.008570862066, 9204.966957103723, 9219.667907983081, 9201.458601228414, 9008.333080014776, 9032.97188569036, 9325.05056835005, 9162.96239371176, 9321.383626648976, 9199.397261239868, 9287.103248368718, 9304.650502181747, 9120.084539905769, 9250.398869948724, 9216.893626623743, 9362.7727638335, 9291.636945042155, 9723.202287388305, 8917.059391427294, 9252.109855327308, 9218.50371293388, 9277.284536344685, 9108.739109104225, 9056.983579474712, 9445.810446372598, 9171.336259955218, 9124.03172061649, 9207.808768997302, 9084.280368928497, 9237.758629150343, 9036.234634500364, 9114.081653787192, 9190.328585092546, 9407.646333580447, 9226.940096547549, 9321.21337608182, 9052.662787088566, 9336.49062114691, 9054.165239295487, 9425.955207984916, 8982.252921851761, 9290.618186798665, 9210.588321890385, 9197.953123607393, 9349.978498704337, 9155.82767648542, 9116.883997438867, 9144.37618267259, 9265.404507743859, 9301.50102467474, 9259.187035649318, 9137.341254601622, 9339.259299087333, 9279.813362119208, 9292.287405790084, 9054.613809092125, 9268.92629501494, 9159.5209006013, 9226.658777291223, 8892.9525801342, 9213.66204579292, 9204.938698683989, 9250.897219198769, 9334.731572991603, 9202.919382311984, 9006.51308889582, 9363.629391180384, 8836.036460908355, 9135.561810947751, 9165.715475881094, 9125.367128800961, 9207.836048343066, 9087.997996009846, 9071.849887300872, 9362.730221201533, 9020.191586621271],
        'vbf': [9689.0, 9664.124548826496, 9711.396033891158, 9701.501136339632, 9675.781925047448, 9464.46694294553, 9886.091433045447, 9580.615703834434, 9778.172173550738, 9650.641053631987, 9738.705521838507, 9700.051387935619, 9682.355091362393, 9657.244326443226, 9733.69196524959, 9752.850849276632, 9628.960245001917, 9822.666199542668, 9615.474191103218, 9793.833530325377, 9636.358916767002, 9709.859360716184, 9666.763359554976, 9712.269830493218, 9683.107895459521, 9698.931061756793, 9666.007732496508, 9713.72400752326, 9668.687603338443, 9726.401625266153, 9660.101126676665, 9688.709851631314, 9688.673473889034, 9685.803606264346, 9692.178377850052, 9695.417000933983, 9696.313986191912, 9679.357133164769, 9694.342838480788, 9701.601502506433, 9691.32995004564, 9651.586528679223, 9699.397008287482, 9656.969986928194, 9719.818371305859, 9705.571106096097, 9644.94431928618, 9718.677510537647, 9673.490390733012, 9662.817760217818, 9627.469262030845, 9688.991497431609, 9643.125193056185, 9618.29218005718, 9634.652305545747, 9604.839138155636, 9619.536474983002, 9616.763928676613, 9609.295766688343, 9631.586334564352, 9611.550719925619, 9622.83783186718, 9637.31895797894, 9591.241383922406, 9757.195046767662, 9533.309561567476, 9582.58949887981, 9666.972104694027, 9654.838859787307, 9604.903657935181, 9627.32871266489, 9601.818665552413, 9626.898557347744, 9607.356769584563, 9532.932044717807, 9777.57192887533, 9667.165066882177, 9588.597055680828, 9630.9953964009, 9612.956456796377, 9618.933476283932, 9580.604821157638, 9653.239192229565, 9532.04916792691, 9625.956049081235, 9615.189809683201, 9599.354231229288, 9628.829709166994, 9605.84863437913, 9624.283183356396, 9609.796584548112, 9624.140772960342, 9622.38576877645, 9615.110544563167, 9551.765532497117, 9703.377247626553, 9541.101314440622, 9651.247080654928, 9377.583024362288, 9632.626094575817, 9565.732777501746, 9497.355501644679, 9546.495050114157, 9631.219799720626, 9627.321130417584, 9787.443253401269, 9527.911154580734, 9557.671631071258, 9658.88898918577, 9593.200811677149, 9646.804646668974, 9458.49544889568, 9553.440305260816, 9522.79063183153, 9469.000608928287, 9546.609962174163, 9769.886348414826, 9494.587472897532, 9629.825609045722, 9695.881068333294, 9787.069163488368, 9641.945037236324, 9780.187202308569, 9737.991635508932, 9607.104378840108, 9195.022292883961, 9083.20705584805, 9406.87099778951, 9582.608386096035, 9589.06467945798, 9520.18675780478, 9724.764374735549, 9768.311740699477, 9493.044181642774, 9460.542840802047, 9436.633867304436, 8799.084563933413, 9742.068214402823, 9591.866220292282, 9438.311721216303, 9550.926867146016, 9630.18721658198, 9627.292914711634, 9080.505147913183, 9419.944820532246, 9680.054719128306, 9474.12026536831, 9497.493424612101, 9396.213215916232, 9765.134494699103, 9642.193527310264, 9666.900087755226, 9465.961914853036, 9607.167522388005, 9425.253876669729, 9767.202551308364, 9588.517401170988, 9524.689858697908, 9511.576375090504, 9701.306127482392, 9219.012548369457, 9650.125638857053, 9664.657936449441, 9434.194954655155, 9714.069179425702, 9781.443495296862, 9737.882325122897, 9569.415148570117, 9362.10237768164, 9485.639468934589, 9529.286847608215, 8957.297715227998, 9658.389532860712, 9560.719681585699, 9726.002207312365, 9453.929949834119, 9658.278173226978, 9582.307995855654, 9841.002355510094, 9306.556998191061, 9283.484317852437, 9416.38955860337, 9424.973172127324, 9598.469576925561, 9555.28564127789, 9384.836523246158, 10076.39939658195, 9860.841006747867, 9500.254766364975, 9497.803177166825, 9507.406106067914, 9641.621998151497, 9537.361544780688, 9182.678046971225, 9783.476717498515],
        }

ARGV = sys.argv[1:]

workspaceDir, = ARGV
#----------------------------------------

# expanded pdf information
pdfInfo = []
# theSum = 0

for pdfFamilyIndex, line in enumerate(pdfMembers):
    for i in range(line[1]):
        pdfInfo.append([ line[0], i, pdfFamilyIndex])

    # theSum += line[1]
    

numSigEventsAfterSelection = {}

# find the workspace files
import glob
for fname in glob.glob(workspaceDir + "/workspace-*.root"):

    basename = os.path.basename(fname)
    print basename
    mo = re.match("workspace-(\d+).root$", basename)
    assert mo

    pdfIndex = int(mo.group(1), 10)

    numSigEventsAfterSelection[pdfIndex] = getSumOfWeightsWorkspace(fname)

# end of loop over files

# pprint(numSigEventsAfterSelection)

#----------
# write this to a spreadsheet
#----------
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

allCats = list(numSigEventsAfterSelection[0].keys())
allCats.sort(key = lambda cat: int(re.match('cat(\d+)$', cat).group(1)))

allProcs = [ 'ggh', 'vbf' ]

numCats = len(allCats)

ws.cell(column = 1, row = 1, value = 'global pdfIndex')
ws.cell(column = 2, row = 1, value = 'pdf family')
ws.cell(column = 3, row = 1, value = 'in-family index')
ws.cell(column = 4, row = 1, value = 'sum weights globe tuples')

firstNumEventsColumn = 6

firstRatioColumn = firstNumEventsColumn + numCats + 1

for procIndex, proc in enumerate(allProcs):

    row = 1
    ws.cell(
        row = row + (4 + numPDFs) * procIndex,
        column = firstNumEventsColumn,
        value = proc)

    for catIndex, cat in enumerate(allCats):
        # row = 2 + (4 + numPDFs) * procIndex
        row = 2 + (4 + numPDFs + 3 * numPDFfamilies) * procIndex

        # normal observable
        ws.cell(
            row = row,
            column = firstNumEventsColumn + catIndex,
            value = cat)

        # delta+ / delta- columns
        col = firstRatioColumn  + numCats + 1 + 2 * catIndex

        ws.cell(
            row = row,
            column = col,
            value = cat)

        for deltaIndex, deltaName in enumerate(['+', '-']):
            ws.cell(
                row = row + 1,
                column = col + deltaIndex,
                value = 'delta' + deltaName)


for procIndex, proc in enumerate(allProcs):

    # names of cells over which the maximum should be taken
    # (per category)
    #
    # first index is category name
    cellsForMax = {}

    for pdfIndex in sorted(numSigEventsAfterSelection):

        pdfFamily = pdfInfo[pdfIndex][0]
        inFamilyIndex = pdfInfo[pdfIndex][1]
        pdfFamilyIndex = pdfInfo[pdfIndex][2]

        row = 4 + pdfIndex + (4 + numPDFs + 3 * numPDFfamilies) * procIndex + 3 * pdfFamilyIndex

        ws.cell(column = 1, row = row, value = pdfIndex)

        if inFamilyIndex == 0:
            nominalPdfRow = row

        ws.cell(column = 2, row = row, value = pdfFamily)
        ws.cell(column = 3, row = row, value = inFamilyIndex)

        #----------
        # sum of weights in globe tuples
        #----------
        sumGlobeWeightsCell = ws.cell(column = 4, row = row, value = sumWeightsGlobeTuples[proc][pdfIndex])

        #----------

        for catIndex, cat in enumerate(allCats):
            cellAfterSelection = ws.cell(
                column = firstNumEventsColumn + catIndex,
                row = row,
                value = numSigEventsAfterSelection[pdfIndex][cat][proc])

            #----------
            # ratio of events after selection to sum of globe tuples
            #----------

            ws.cell(
                column = firstRatioColumn + catIndex,
                row = row,
                value = "=%s / %s" % (cellAfterSelection.coordinate, sumGlobeWeightsCell.coordinate)
                )

        #----------
        # delta columns
        #----------
        
        # find the number of pdfs in this family
        familySize = [ line[1] for line in pdfMembers if line[0] == pdfFamily]
        assert len(familySize) == 1
        familySize = familySize[0]

        #----------
        
        if pdfFamily in pdfsWithUpDown:
            if inFamilyIndex == 0:
                addPerPDFfamilyUpDownSummary(row, cellsForMax)

            if inFamilyIndex > 0 and inFamilyIndex % 2 == 0:

                for catIndex, cat in enumerate(allCats):

                    # for number of events
                    # col = firstNumEventsColumn + numCats + 1 + 2 * catIndex
                    # for 'acceptance'
                    col = firstRatioColumn + numCats + 1 + 2 * catIndex

                    origCol = openpyxl.cell.get_column_letter(firstRatioColumn + catIndex)

                    ws.cell(
                        column = col,
                        row = row,
                        value = '=MAX(%s%d - %s$%d, %s%d - %s$%d, 0)' % (origCol, row - 1,
                                                                         origCol, nominalPdfRow,
                                                                         origCol, row,
                                                                         origCol, nominalPdfRow)
                        )

                    

                    ws.cell(
                        column = col + 1,
                        row = row,
                        value = '=MAX(%s$%d - %s%d, %s$%d - %s%d, 0)' % (origCol, nominalPdfRow,
                                                                         origCol, row - 1,
                                                                         origCol, nominalPdfRow,
                                                                         origCol, row)
                        )

        else:
            # this is NOT a PDF with up/down
            # we take the rms
            # also add a row with the average of our observable
            # (as opposed of the observable of the average)

            if inFamilyIndex == 0:
                    addPerPDFfamilyNNPDFsummary(row, cellsForMax)

            else:
                pass
                # for catIndex, cat in enumerate(allCats):
                # 
                #     col = 5 + numCats + 1 + 2 * catIndex
                # 
                #     origCol = openpyxl.cell.get_column_letter(5 + catIndex)
                # 
                #     ws.cell(
                #         column = col,
                #         row = row,
                #         value = '=(%s%d - %s$%d, %s%d - %s$%d, 0)' % (origCol, row - 1,
                #                                                          origCol, nominalPdfRow,
                #                                                          origCol, row,
                #                                                          origCol, nominalPdfRow)
                #         )
                
    # end of loop over PDFs

    # put the cells for the maximum over all PDF families

    # row = 4 + (4 + numPDFs + 3 * numPDFfamilies) * len(allProcs)
    row = 4 + numPDFs + (4 + numPDFs + 3 * numPDFfamilies) * procIndex + 3 * (numPDFfamilies - 1) + 2

    addMaxOverPDFfamilies(row,
                          # row where the nominal values are (PDF index = 0)
                          4 + 0 + (4 + numPDFs + 3 * numPDFfamilies) * procIndex,
                          cellsForMax
                          )

# end of loop over processes

wb.save(filename = "pdf-uncert.xlsx")


#----------------------------------------

### # pprint(numSigEvents); sys.exit(1)
### 
### pdfs = numSigEvents.keys()
### pdfs.remove('central')
### pdfs.sort()
### 
### # assert len(pdfs) == 2
### 
### pdfNums = set()
### for pdf in pdfs:
###     mo = re.match("(up|down)(\d+)$", pdf)
###     assert mo
###     # we keep this as an integer on purpose
###     pdfNums.add(mo.group(2))
### 
### 
### pdfNums = sorted(pdfNums)
### print "found",pdfNums,"pdf directions"
### 
### # first index is category
### # second index is signal process
### # third index is the pdf number
### relDeviations = {}
### 
### for proc in allProcs:
### 
###     #----------
###     # calculate sum of events over all categories
###     #----------
### 
###     sumEventsUp = {}
###     sumEventsDown = {}
### 
###     for pdfNum in pdfNums:
###         sumEventsUp[pdfNum]   = sum([numSigEvents['up'   + pdfNum][cat][proc] for cat in allCats])
###         sumEventsDown[pdfNum] = sum([numSigEvents['down' + pdfNum][cat][proc] for cat in allCats])
### 
###     sumEventsCentral = sum([numSigEvents['central'][cat][proc] for cat in allCats])
### 
###     #----------
###     
###     for cat in allCats:
### 
###         nom = numSigEvents['central'][cat][proc]
### 
###         sumSqRel = 0
### 
###         for pdfNum in pdfNums:
###             up = numSigEvents['up' + pdfNum][cat][proc]
###             down = numSigEvents['down' + pdfNum][cat][proc]
### 
###             if averageDeviations:
###                 # take out any average change in normalization
###                 up   *= sumEventsCentral / float(sumEventsUp[pdfNum])
###                 down *= sumEventsCentral / float(sumEventsDown[pdfNum])
### 
###             # check that up and down are on opposite sides of nominal
### 
###             if not ((up >= nom and nom >= down) or (up <= nom and nom <= down)):
###                 print "WARNING: up/down are NOT on opposite sides of nominal for",cat,proc,": up=",up,"nom=",nom,"down=",down
### 
###                 # take the maximum deviation as uncertainty, with the sign determined by the larger deviation
###                 relup = (up - nom) / nom
###                 reldown = (nom - down) / nom
### 
###                 if abs(relup) > abs(reldown):
###                     rel = relup
###                 else:
###                     rel = reldown
###             else:
###                 # deviations are on both sides
###                 # take (up - down) / (2 * nominal)
###                 rel = (up - down) / (2.0 * nom)
### 
###             sumSqRel += rel * rel
### 
###             relDeviations.setdefault(cat,{}).setdefault(proc,{})[pdfNum] = rel
### 
###         # end of loop over pdf numbers
### 
### pprint(relDeviations);# sys.exit(1)
### 
### 
### print "         "," ".join([ " %5s" % cat for cat in allCats])
### for proc in allProcs:
###     print "%-3s" % proc
### 
###     sums = [ 0] * len(allCats)
### 
###     for pdfNum in pdfNums:
###         parts = [ "%+5.1f%%" % (relDeviations[cat][proc][pdfNum] * 100) for cat in allCats ]
### 
###         sums = [ x + relDeviations[cat][proc][pdfNum]**2 for cat,x in zip(allCats,sums) ]
### 
###         print "       %02s" % pdfNum," ".join(parts)
### 
###     # print the quadratic sum:
###     sums = [ math.sqrt(x) for x in sums ]
###     
###     parts = [ "%5.1f%%" % (x * 100) for x in sums ]
###     
###     print "      %03s" % "sum"," ".join(parts)        
